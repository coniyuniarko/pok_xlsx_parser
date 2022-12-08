import openpyxl
import re


workbook = openpyxl.load_workbook("fa_detail.xlsx")

worksheet = workbook.active

program = ""
kegiatan = ""
output = ""
suboutput = ""
komponen = ""
subkomponen = ""
akun = ""

for row in range(9, worksheet.max_row):

	column_b = str(worksheet.cell(row=row, column=2).value)
	column_c = str(worksheet.cell(row=row, column=3).value)
	column_d = str(worksheet.cell(row=row, column=4).value)
	column_e = str(worksheet.cell(row=row, column=5).value)
	column_f = str(worksheet.cell(row=row, column=6).value)
	column_g = str(worksheet.cell(row=row, column=7).value)
	column_h = str(worksheet.cell(row=row, column=8).value)
	column_i = str(worksheet.cell(row=row, column=9).value)
	column_j = str(worksheet.cell(row=row, column=10).value)
	column_k = str(worksheet.cell(row=row, column=11).value)
	column_l = str(worksheet.cell(row=row, column=12).value)
	column_m = str(worksheet.cell(row=row, column=13).value)
	column_p = worksheet.cell(row=row, column=16).value

	if column_b.startswith("*Lock"):
		continue

	program_cek = column_b + " " + column_d
	match = re.search("[A-Z]{2}\s[\w]", program_cek)
	if match:
		program = program_cek
		continue

	match = re.search("[^A-Z\.\D*]", column_b)
	if match:
		kegiatan = column_b.split(".")[1] + " " + column_i
		continue
	
	match = re.search("^[A-Z]{3}$", column_c)
	if match:
		output = column_c + " " + column_g
		continue

	match = re.search("^[A-Z]{3}\.[\d]{3}$", column_c)
	if match:
		suboutput = column_c.split(".")[1] + " " + column_k
		continue
	
	match = re.search("^[\d]{3}$", column_e)
	if match:
		komponen = column_e + " " + column_j
		continue
	
	match = re.search("^[\d]{3}\.[\d]{1}[A-Z]{1}$", column_f)
	if match:
		subkomponen = column_f.split(".")[1] + " " + column_l
		continue
	
	match = re.search("^5[\d]{5}$", column_h)
	if match:
		akun = column_h + " " + column_m
		print("Program : ", program)
		print("Kegiatan : ", kegiatan)
		print("Output : ", output)
		print("Suboutput : ", suboutput)
		print("Komponen : ", komponen)
		print("SubKomponen : ", subkomponen)
		print("Akun : ", akun)
		print("Pagu : ", int(column_p))
		print()
		continue
